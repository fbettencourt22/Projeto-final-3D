import io
import json
from pathlib import Path
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
import unicodedata

from django.contrib import messages
from django.contrib.auth import logout
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, HttpResponseForbidden
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.utils.http import url_has_allowed_host_and_scheme
from django.db.models import Q

from .forms import FilamentTypeForm, InventoryQuantityForm, PieceImportForm, PrintJobForm
from .models import FilamentType, InventoryItem, PrintJob

VALOR_KWH = Decimal("0.158")
CONSUMO_W = Decimal("140")
CUSTO_MAO_OBRA = Decimal("20")
DESPERDICIO_FILAMENTO = Decimal("0.10")
CUSTO_MAQUINA_HORA = Decimal("0.20")

IMPORT_COLUMNS = [
    "piece_name",
    "filament_price_per_kg",
    "filament_weight_g",
    "print_time_hours",
    "labour_time_minutes",
    "margin_percentage",
]


def to_currency(value: Decimal) -> Decimal:
    return value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


def calculate_print_job(data: dict) -> dict:
    filament_price_per_kg = Decimal(data["filament_price_per_kg"])
    filament_weight_g = Decimal(data["filament_weight_g"])
    print_time_hours = Decimal(data["print_time_hours"])
    labour_time_minutes = Decimal(data["labour_time_minutes"])
    margin_percentage = Decimal(data["margin_percentage"])

    cost_filament = (filament_price_per_kg / Decimal("1000")) * filament_weight_g
    cost_filament *= Decimal("1") + DESPERDICIO_FILAMENTO

    consumption_kwh = (CONSUMO_W * print_time_hours) / Decimal("1000")
    cost_energy = consumption_kwh * VALOR_KWH

    cost_labour = (labour_time_minutes / Decimal("60")) * CUSTO_MAO_OBRA
    cost_machine = print_time_hours * CUSTO_MAQUINA_HORA

    cost_total = cost_filament + cost_energy + cost_labour + cost_machine
    price_final = cost_total / (Decimal("1") - (margin_percentage / Decimal("100")))

    return {
        "cost_filament": to_currency(cost_filament),
        "cost_energy": to_currency(cost_energy),
        "cost_labour": to_currency(cost_labour),
        "cost_machine": to_currency(cost_machine),
        "cost_total": to_currency(cost_total),
        "price_final": to_currency(price_final),
        "consumption_kwh": consumption_kwh.quantize(
            Decimal("0.0001"), rounding=ROUND_HALF_UP
        ),
    }




def get_piece_initial_data(piece: PrintJob, user) -> dict:
    initial = {
        "piece_name": piece.name,
        "filament_type": piece.filament_type,
        "filament_weight_g": piece.filament_weight_g,
        "print_time_hours": piece.print_time_hours,
        "labour_time_minutes": piece.labour_time_minutes,
        "margin_percentage": piece.margin_percentage,
    }
    if initial["filament_type"] is None:
        filament_qs = FilamentType.objects.all().order_by("name")
        if user and not getattr(user, "is_superuser", False):
            filament_qs = filament_qs.filter(user=user)
        guess = filament_qs.filter(price_per_kg=piece.filament_price_per_kg).first()
        if guess is not None:
            initial["filament_type"] = guess
    return initial


def serialize_piece_edit_payload(piece: PrintJob, user) -> str:
    initial = get_piece_initial_data(piece, user)
    payload = {
        "pk": piece.pk,
        "piece_name": initial["piece_name"] or "",
        "filament_type": "",
        "filament_weight_g": "",
        "print_time_hours": "",
        "labour_time_minutes": "",
        "margin_percentage": "",
        "label": piece.name or f"Peca #{piece.pk}",
    }
    filament = initial["filament_type"]
    if filament is not None:
        payload["filament_type"] = str(filament.pk)
    for key in (
        "filament_weight_g",
        "print_time_hours",
        "labour_time_minutes",
        "margin_percentage",
    ):
        value = initial[key]
        payload[key] = "" if value is None else str(value)
    return json.dumps(payload, ensure_ascii=False)


def get_filament_label(filament: FilamentType) -> str:
    color = (filament.color or '').strip()
    if color:
        return f"{filament.name} ({color})"
    return filament.name


def serialize_filament_edit_payload(filament: FilamentType) -> str:
    payload = {
        "pk": filament.pk,
        "name": filament.name or "",
        "color": filament.color or "",
        "price_per_kg": "" if filament.price_per_kg is None else str(filament.price_per_kg),
        "weight_kg": "" if filament.weight_kg is None else str(filament.weight_kg),
        "label": get_filament_label(filament),
    }
    return json.dumps(payload, ensure_ascii=False)


def serialize_inventory_item_edit_payload(item: InventoryItem) -> str:
    label_source = item.piece_name or (item.print_job.name if getattr(item, 'print_job', None) and item.print_job.name else '')
    payload = {
        "pk": item.pk,
        "quantity": str(item.quantity or 0),
        "label": label_source or f"Item #{item.pk}",
    }
    return json.dumps(payload, ensure_ascii=False)

def serialize_inventory_add_payload(piece: PrintJob) -> str:
    payload = {
        "pk": piece.pk,
        "quantity": "1",
        "label": piece.name or f"Peca #{piece.pk}",
    }
    return json.dumps(payload, ensure_ascii=False)


def add_piece_to_inventory(user, piece: PrintJob, quantity: int):
    piece_label = piece.name or f"Peca #{piece.pk}"
    item, created = InventoryItem.objects.get_or_create(
        user=user,
        print_job=piece,
        defaults={"quantity": quantity, "piece_name": piece_label},
    )
    if not created:
        item.quantity += quantity
        item.piece_name = piece_label
        item.save(update_fields=["quantity", "piece_name", "updated_at"])
    return item, created, piece_label



def update_piece_from_form(piece: PrintJob, cleaned_data: dict, request_user) -> None:
    cleaned_values = cleaned_data.copy()
    filament = cleaned_values["filament_type"]
    cleaned_values["filament_price_per_kg"] = filament.price_per_kg
    result = calculate_print_job(cleaned_values)

    piece.name = cleaned_values.get("piece_name", "")
    piece.filament_type = filament
    piece.filament_price_per_kg = cleaned_values["filament_price_per_kg"]
    piece.filament_weight_g = cleaned_values["filament_weight_g"]
    piece.print_time_hours = cleaned_values["print_time_hours"]
    piece.labour_time_minutes = cleaned_values["labour_time_minutes"]
    piece.margin_percentage = cleaned_values["margin_percentage"]
    piece.cost_filament = result["cost_filament"]
    piece.cost_energy = result["cost_energy"]
    piece.cost_labour = result["cost_labour"]
    piece.cost_machine = result["cost_machine"]
    piece.cost_total = result["cost_total"]
    piece.price_final = result["price_final"]
    piece.consumption_kwh = result["consumption_kwh"]
    if piece.user is None:
        piece.user = request_user
    piece.save()


def get_safe_redirect(request, default_url: str) -> str:
    next_url = request.POST.get("next") or request.GET.get("next")
    if next_url and url_has_allowed_host_and_scheme(
        next_url,
        allowed_hosts={request.get_host()},
        require_https=request.is_secure(),
    ):
        return next_url
    return default_url


def normalize_text(value: str) -> str:
    if not value:
        return ''
    normalized = unicodedata.normalize('NFKD', value)
    stripped = ''.join(ch for ch in normalized if not unicodedata.combining(ch))
    return stripped.lower()


def parse_decimal(value):
    """Converte o valor em Decimal, aceitando strings com vArgula."""
    if isinstance(value, Decimal):
        return value
    if value is None:
        raise ValueError("valor em falta")
    value_str = str(value).strip()
    if not value_str:
        raise ValueError("valor em falta")
    value_str = value_str.replace(",", ".")
    try:
        return Decimal(value_str)
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"valor invAlido: {value}") from exc





def piece_permission_check(user, piece: PrintJob):
    if user.is_superuser:
        return True
    return piece.user == user


@login_required
def logout_view(request):
    logout(request)
    return redirect("login")


@login_required
def dashboard_view(request):
    links = [
        {'url': 'calculator', 'label': 'Calculadora', 'description': 'Calculadora de custos de impressão 3D.'},
        {'url': 'pieces_list', 'label': 'Pe\u00e7as inseridas', 'description': 'Consulte, edite e exporte os seus cálculos anteriores.'},
        {'url': 'piece_import', 'label': 'Importar pe\u00e7as', 'description': 'Carregue um ficheiro Excel para criar pe\u00e7as em massa.'},
        {'url': 'inventory', 'label': 'Inventário', 'description': 'Gerir filamentos e pe\u00e7as disponíveis.'},
    ]
    return render(request, 'core/dashboard.html', {'links': links})


@login_required
def inventory_view(request):
    active_tab = request.GET.get('tab', 'filaments')
    if active_tab not in {'filaments', 'pieces'}:
        active_tab = 'filaments'

    filament_base_qs = FilamentType.objects.all() if request.user.is_superuser else FilamentType.objects.filter(user=request.user)
    filaments = list(filament_base_qs.order_by('name'))

    inventory_items_qs = (
        InventoryItem.objects.filter(user=request.user)
        .select_related('print_job')
        .order_by('piece_name')
    )
    inventory_items_list = list(inventory_items_qs)

    pieces_search = request.GET.get('pieces_search', '').strip()

    if pieces_search:
        norm_search = normalize_text(pieces_search)
        inventory_items_list = [
            item
            for item in inventory_items_list
            if norm_search in normalize_text(item.piece_name or '')
            or (item.print_job and norm_search in normalize_text(item.print_job.name or ''))
            or (pieces_search.isdigit() and item.print_job and pieces_search == str(item.print_job.pk))
        ]
        active_tab = 'pieces'

    def build_next_url(tab_name: str, exclude_key: str) -> str:
        query_params = request.GET.copy()
        query_params['tab'] = tab_name
        query_params.pop(exclude_key, None)
        if query_params:
            return f"{request.path}?{query_params.urlencode()}"
        return request.path

    filament_form = FilamentTypeForm()
    filament_edit_form = FilamentTypeForm()
    filament_edit_open_pk = None
    filament_edit_next_url = request.get_full_path()

    inventory_item_edit_form = InventoryQuantityForm()
    inventory_item_edit_open_pk = None
    inventory_item_edit_next_url = request.get_full_path()

    action = request.POST.get('action') if request.method == 'POST' else None
    if action == 'add_filament':
        filament_form = FilamentTypeForm(request.POST)
        if filament_form.is_valid():
            filament = filament_form.save(commit=False)
            filament.user = request.user
            filament.save()
            messages.success(request, 'Filamento guardado no invent\u00e1rio.')
            return redirect(f"{reverse('inventory')}?tab=filaments")
        active_tab = 'filaments'
    elif action == 'edit_filament':
        active_tab = 'filaments'
        filament_pk = request.POST.get('filament_id')
        filament = get_object_or_404(filament_base_qs, pk=filament_pk)
        filament_edit_form = FilamentTypeForm(request.POST, instance=filament)
        if filament_edit_form.is_valid():
            filament_edit_form.save()
            messages.success(request, 'Filamento atualizado.')
            target = get_safe_redirect(request, f"{reverse('inventory')}?tab=filaments")
            return redirect(target)
        filament_edit_open_pk = str(filament.pk)
        filament_edit_next_url = request.POST.get('next') or build_next_url('filaments', 'open_filament_edit')
    elif action == 'edit_inventory_item':
        active_tab = 'pieces'
        item_pk = request.POST.get('item_id')
        item = get_object_or_404(inventory_items_qs, pk=item_pk)
        inventory_item_edit_form = InventoryQuantityForm(request.POST)
        if inventory_item_edit_form.is_valid():
            item.quantity = inventory_item_edit_form.cleaned_data['quantity']
            item.save(update_fields=['quantity', 'updated_at'])
            messages.success(request, 'Invent\u00e1rio atualizado.')
            target = get_safe_redirect(request, f"{reverse('inventory')}?tab=pieces")
            return redirect(target)
        inventory_item_edit_open_pk = str(item.pk)
        inventory_item_edit_next_url = request.POST.get('next') or build_next_url('pieces', 'open_inventory_item_edit')

    if action != 'edit_filament':
        filament_edit_next_url = build_next_url('filaments', 'open_filament_edit')
    if action != 'edit_inventory_item':
        inventory_item_edit_next_url = build_next_url('pieces', 'open_inventory_item_edit')

    if filament_edit_open_pk is None:
        open_filament_pk = request.GET.get('open_filament_edit')
        if open_filament_pk and filament_base_qs.filter(pk=open_filament_pk).exists():
            filament_edit_open_pk = open_filament_pk
            active_tab = 'filaments'

    if inventory_item_edit_open_pk is None:
        open_item_pk = request.GET.get('open_inventory_item_edit')
        if open_item_pk and inventory_items_qs.filter(pk=open_item_pk).exists():
            inventory_item_edit_open_pk = open_item_pk
            active_tab = 'pieces'

    for filament in filaments:
        filament.edit_payload = serialize_filament_edit_payload(filament)
        filament.edit_label = get_filament_label(filament)

    for item in inventory_items_list:
        item.edit_payload = serialize_inventory_item_edit_payload(item)
        item.edit_label = item.piece_name or f"Item #{item.pk}"

    return render(
        request,
        'core/inventory.html',
        {
            'filament_form': filament_form,
            'filaments': filaments,
            'inventory_items': inventory_items_list,
            'active_tab': active_tab,
            'pieces_search': pieces_search,
            'filament_edit_form': filament_edit_form,
            'filament_edit_open_pk': filament_edit_open_pk,
            'filament_edit_next_url': filament_edit_next_url,
            'inventory_item_edit_form': inventory_item_edit_form,
            'inventory_item_edit_open_pk': inventory_item_edit_open_pk,
            'inventory_item_edit_next_url': inventory_item_edit_next_url,
        },
    )


@login_required
def inventory_add_piece_view(request, pk):
    piece = get_object_or_404(PrintJob, pk=pk)
    if not piece_permission_check(request.user, piece):
        return HttpResponseForbidden('Nao autorizado')

    form = InventoryQuantityForm()
    if request.method == 'POST':
        form = InventoryQuantityForm(request.POST)
        if form.is_valid():
            quantity = form.cleaned_data['quantity']
            _, created, piece_label = add_piece_to_inventory(request.user, piece, quantity)
            message = 'Peca adicionada ao inventario.' if created else 'Quantidade atualizada no inventario.'
            messages.success(request, message)
            return redirect(f"{reverse('inventory')}?tab=pieces")

    return render(
        request,
        'core/inventory_add_piece.html',
        {
            'form': form,
            'piece': piece,
        },
    )


@login_required
def inventory_filament_edit_view(request, pk):
    qs = FilamentType.objects.all() if request.user.is_superuser else FilamentType.objects.filter(user=request.user)
    filament = get_object_or_404(qs, pk=pk)
    target = f"{reverse('inventory')}?tab=filaments&open_filament_edit={filament.pk}"
    return redirect(target)


@login_required
def inventory_filament_delete_view(request, pk):
    qs = FilamentType.objects.all() if request.user.is_superuser else FilamentType.objects.filter(user=request.user)
    filament = get_object_or_404(qs, pk=pk)

    if request.method == 'POST':
        filament.delete()
        messages.success(request, 'Filamento removido do inventário.')
        return redirect(f"{reverse('inventory')}?tab=filaments")

    return render(
        request,
        'core/inventory_filament_confirm_delete.html',
        {
            'filament': filament,
        },
    )


@login_required
def inventory_item_edit_view(request, pk):
    qs = InventoryItem.objects.select_related('print_job')
    if not request.user.is_superuser:
        qs = qs.filter(user=request.user)
    item = get_object_or_404(qs, pk=pk)
    target = f"{reverse('inventory')}?tab=pieces&open_inventory_item_edit={item.pk}"
    return redirect(target)


@login_required
def inventory_item_delete_view(request, pk):
    qs = InventoryItem.objects.all() if request.user.is_superuser else InventoryItem.objects.filter(user=request.user)
    item = get_object_or_404(qs, pk=pk)

    if request.method == 'POST':
        item.delete()
        messages.success(request, 'Pe\u00e7a removida do invent\u00e1rio.')
        return redirect(f"{reverse('inventory')}?tab=pieces")

    return render(
        request,
        'core/inventory_item_confirm_delete.html',
        {
            'item': item,
        },
    )


@login_required
def calculator_view(request):
    result = None

    form = PrintJobForm(user=request.user)
    piece_edit_form = PrintJobForm(user=request.user)
    piece_edit_open_pk = None

    inventory_add_form = InventoryQuantityForm()
    inventory_add_open_pk = None
    inventory_add_next_url = request.get_full_path()

    action = request.POST.get("action") if request.method == "POST" else None
    if request.method == "POST":
        if action == "add_to_inventory":
            inventory_add_form = InventoryQuantityForm(request.POST)
            piece_id = request.POST.get("piece_id")
            piece = get_object_or_404(PrintJob.objects.select_related("user"), pk=piece_id)
            if not piece_permission_check(request.user, piece):
                return HttpResponseForbidden("Sem permissao.")
            if inventory_add_form.is_valid():
                quantity = inventory_add_form.cleaned_data["quantity"]
                _, _, piece_label = add_piece_to_inventory(request.user, piece, quantity)
                messages.success(request, f"Peca '{piece_label}' enviada para o inventario.")
                return redirect(get_safe_redirect(request, request.POST.get("next") or reverse("calculator")))
            inventory_add_open_pk = str(piece.pk)
            inventory_add_next_url = request.POST.get("next") or request.get_full_path()
        else:
            piece_id = request.POST.get("piece_id")
            if piece_id:
                piece = get_object_or_404(PrintJob.objects.select_related("user"), pk=piece_id)
                if not piece_permission_check(request.user, piece):
                    return HttpResponseForbidden("Sem permissao.")
                piece_edit_form = PrintJobForm(request.POST, user=request.user, existing_piece=piece)
                if piece_edit_form.is_valid():
                    update_piece_from_form(piece, piece_edit_form.cleaned_data, request.user)
                    messages.success(request, "Peca atualizada.")
                    return redirect(get_safe_redirect(request, reverse("calculator")))
                piece_edit_open_pk = str(piece.pk)
            else:
                form = PrintJobForm(request.POST, user=request.user)
                if form.is_valid():
                    cleaned = form.cleaned_data
                    filament = cleaned["filament_type"]
                    cleaned["filament_price_per_kg"] = filament.price_per_kg
                    result = calculate_print_job(cleaned)
                    result["filament_name"] = filament.name
                    result["filament_color"] = filament.color
                    result["filament_price_per_kg"] = to_currency(filament.price_per_kg)

                    print_job = PrintJob.objects.create(
                        user=request.user,
                        name=cleaned.get("piece_name", ""),
                        filament_type=filament,
                        filament_price_per_kg=cleaned["filament_price_per_kg"],
                        filament_weight_g=cleaned["filament_weight_g"],
                        print_time_hours=cleaned["print_time_hours"],
                        labour_time_minutes=cleaned["labour_time_minutes"],
                        margin_percentage=cleaned["margin_percentage"],
                        cost_filament=result["cost_filament"],
                        cost_energy=result["cost_energy"],
                        cost_labour=result["cost_labour"],
                        cost_machine=result["cost_machine"],
                        cost_total=result["cost_total"],
                        price_final=result["price_final"],
                        consumption_kwh=result["consumption_kwh"],
                    )

                    piece_label = print_job.name or f"#{print_job.pk}"
                    messages.success(
                        request,
                        f"Peca '{piece_label}' calculada e guardada com sucesso.",
                    )

                    result["piece_name"] = print_job.name or f"Peca #{print_job.pk}"
                    result["created_at"] = print_job.created_at

                    form = PrintJobForm(user=request.user)

    pieces_qs = PrintJob.objects.select_related("user")
    if request.user.is_superuser:
        pieces_queryset = pieces_qs.exclude(inventory_records__isnull=False).distinct()
    else:
        pieces_queryset = (
            pieces_qs
            .filter(user=request.user)
            .exclude(inventory_records__user=request.user)
            .distinct()
        )

    pieces_list = list(pieces_queryset)
    for piece in pieces_list:
        piece.edit_payload = serialize_piece_edit_payload(piece, request.user)
        piece.edit_label = piece.name or f"Peca #{piece.pk}"
        piece.add_payload = serialize_inventory_add_payload(piece)

    if piece_edit_open_pk is None:
        open_edit_pk = request.GET.get("open_edit")
        if open_edit_pk:
            try:
                piece = pieces_qs.get(pk=open_edit_pk)
            except PrintJob.DoesNotExist:
                piece_edit_open_pk = None
            else:
                if piece_permission_check(request.user, piece):
                    piece_edit_open_pk = str(piece.pk)

    has_filaments = form.fields["filament_type"].queryset.exists()

    if request.method == "POST" and request.POST.get("piece_id") and action != "add_to_inventory":
        piece_edit_next_url = request.POST.get("next") or request.get_full_path()
    else:
        piece_edit_next_url = request.get_full_path()

    if not (request.method == "POST" and action == "add_to_inventory"):
        params = request.GET.copy()
        params.pop("open_edit", None)
        inventory_add_next_url = request.path
        if params:
            inventory_add_next_url = f"{inventory_add_next_url}?{params.urlencode()}"

    context = {
        "form": form,
        "result": result,
        "pieces": pieces_list,
        "has_filaments": has_filaments,
        "constants": {
            "VALOR_KWH": VALOR_KWH,
            "CONSUMO_W": CONSUMO_W,
            "CUSTO_MAO_OBRA": CUSTO_MAO_OBRA,
            "DESPERDICIO_FILAMENTO": DESPERDICIO_FILAMENTO,
            "CUSTO_MAQUINA_HORA": CUSTO_MAQUINA_HORA,
        },
        "piece_edit_form": piece_edit_form,
        "piece_edit_open_pk": piece_edit_open_pk,
        "piece_edit_next_url": piece_edit_next_url,
        "inventory_add_form": inventory_add_form,
        "inventory_add_open_pk": inventory_add_open_pk,
        "inventory_add_next_url": inventory_add_next_url,
        "inventory_add_action": "calculator",
    }
    return render(request, "core/calculator.html", context)


@login_required
def pieces_list_view(request):
    pieces_qs = PrintJob.objects.select_related("user")
    if request.user.is_superuser:
        base_queryset = pieces_qs.exclude(inventory_records__isnull=False).distinct()
    else:
        base_queryset = (
            pieces_qs
            .filter(user=request.user)
            .exclude(inventory_records__user=request.user)
            .distinct()
        )

    edit_form = PrintJobForm(user=request.user)
    active_piece_pk = None

    inventory_add_form = InventoryQuantityForm()
    inventory_add_open_pk = None
    inventory_add_next_url = request.get_full_path()

    if request.method == "POST":
        action = request.POST.get("action")
        if action == "add_to_inventory":
            inventory_add_form = InventoryQuantityForm(request.POST)
            piece_id = request.POST.get("piece_id")
            piece = get_object_or_404(pieces_qs, pk=piece_id)
            if not piece_permission_check(request.user, piece):
                return HttpResponseForbidden("Sem permissao.")
            if inventory_add_form.is_valid():
                quantity = inventory_add_form.cleaned_data["quantity"]
                _, _, piece_label = add_piece_to_inventory(request.user, piece, quantity)
                messages.success(request, f"Peca '{piece_label}' enviada para o inventario.")
                redirect_target = get_safe_redirect(request, request.POST.get("next") or reverse("pieces_list"))
                return redirect(redirect_target)
            inventory_add_open_pk = str(piece.pk)
            inventory_add_next_url = request.POST.get("next") or request.get_full_path()
        elif request.POST.get("piece_id"):
            piece = get_object_or_404(pieces_qs, pk=request.POST.get("piece_id"))
            if not piece_permission_check(request.user, piece):
                return HttpResponseForbidden("Sem permissao.")
            edit_form = PrintJobForm(request.POST, user=request.user, existing_piece=piece)
            if edit_form.is_valid():
                update_piece_from_form(piece, edit_form.cleaned_data, request.user)
                messages.success(request, "Peca atualizada.")
                redirect_target = get_safe_redirect(request, reverse("pieces_list"))
                return redirect(redirect_target)
            active_piece_pk = str(piece.pk)

    search_query = request.GET.get("search", "").strip()
    queryset = base_queryset
    if search_query:
        filters = Q(name__icontains=search_query)
        if search_query.isdigit():
            filters |= Q(pk=int(search_query))
        queryset = queryset.filter(filters).distinct()

    pieces_list = list(queryset)
    if search_query:
        norm_query = normalize_text(search_query)
        pieces_list = [
            piece
            for piece in pieces_list
            if norm_query in normalize_text(piece.name or "")
            or norm_query in normalize_text(str(piece.pk))
        ]

    for piece in pieces_list:
        piece.edit_payload = serialize_piece_edit_payload(piece, request.user)
        piece.edit_label = piece.name or f"Peca #{piece.pk}"
        piece.add_payload = serialize_inventory_add_payload(piece)

    if active_piece_pk is None:
        open_edit_pk = request.GET.get("open_edit")
        if open_edit_pk:
            try:
                piece = pieces_qs.get(pk=open_edit_pk)
            except PrintJob.DoesNotExist:
                active_piece_pk = None
            else:
                if piece_permission_check(request.user, piece):
                    active_piece_pk = str(piece.pk)

    if request.method == "POST" and request.POST.get("piece_id") and request.POST.get("action") != "add_to_inventory":
        piece_edit_next_url = request.POST.get("next") or reverse("pieces_list")
    else:
        query_params = request.GET.copy()
        query_params.pop("open_edit", None)
        piece_edit_next_url = request.path
        if query_params:
            piece_edit_next_url = f"{piece_edit_next_url}?{query_params.urlencode()}"

    if not (request.method == "POST" and request.POST.get("action") == "add_to_inventory"):
        query_params = request.GET.copy()
        query_params.pop("open_edit", None)
        inventory_add_next_url = request.path
        if query_params:
            inventory_add_next_url = f"{inventory_add_next_url}?{query_params.urlencode()}"

    context = {
        "pieces": pieces_list,
        "search_query": search_query,
        "piece_edit_form": edit_form,
        "piece_edit_open_pk": active_piece_pk,
        "piece_edit_next_url": piece_edit_next_url,
        "inventory_add_form": inventory_add_form,
        "inventory_add_open_pk": inventory_add_open_pk,
        "inventory_add_next_url": inventory_add_next_url,
        "inventory_add_action": "pieces_list",
    }
    return render(request, "core/pieces_list.html", context)

@login_required
def piece_edit_view(request, pk: int):
    piece = get_object_or_404(PrintJob.objects.select_related("user"), pk=pk)
    if not piece_permission_check(request.user, piece):
        return HttpResponseForbidden("Sem permissao.")
    target = f"{reverse('pieces_list')}?open_edit={piece.pk}"
    return redirect(target)

@login_required
def piece_delete_view(request, pk: int):
    piece = get_object_or_404(PrintJob.objects.select_related("user"), pk=pk)
    if not piece_permission_check(request.user, piece):
        return HttpResponseForbidden("Sem permissão.")

    if request.method == "POST":
        piece.delete()
        return redirect("pieces_list")

    return render(request, "core/piece_confirm_delete.html", {"piece": piece})




@login_required
def piece_export_view(request):
    pieces = PrintJob.objects.select_related("user")
    if not request.user.is_superuser:
        pieces = pieces.filter(user=request.user)

    try:
        from openpyxl import Workbook  # type: ignore
    except ImportError:
        messages.error(
            request,
            "Suporte a Excel indisponAvel (biblioteca openpyxl nAo instalada).",
        )
        return redirect("pieces_list")

    wb = Workbook()
    ws = wb.active
    ws.title = "Pe\u00e7as"
    headers = [
        "piece_name",
        "filament_price_per_kg",
        "filament_weight_g",
        "print_time_hours",
        "labour_time_minutes",
        "margin_percentage",
        "cost_filament",
        "cost_energy",
        "cost_labour",
        "cost_machine",
        "cost_total",
        "price_final",
        "consumption_kwh",
        "created_at",
        "owner",
    ]
    ws.append(headers)

    for piece in pieces:
        created_at = piece.created_at
        if created_at is not None:
            if timezone.is_aware(created_at):
                created_at = timezone.localtime(created_at)
            created_at = created_at.replace(tzinfo=None)
        ws.append(
            [
                piece.name or "",
                float(piece.filament_price_per_kg),
                float(piece.filament_weight_g),
                float(piece.print_time_hours),
                float(piece.labour_time_minutes),
                float(piece.margin_percentage),
                float(piece.cost_filament),
                float(piece.cost_energy),
                float(piece.cost_labour),
                float(piece.cost_machine),
                float(piece.cost_total),
                float(piece.price_final),
                float(piece.consumption_kwh),
                created_at,
                piece.user.get_username() if piece.user else "",
            ]
        )

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
    response[
        "Content-Disposition"
    ] = f'attachment; filename="pe\u00e7as_{timestamp}.xlsx"'

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response.write(buffer.getvalue())
    return response



@login_required
def piece_import_view(request):
    form = PieceImportForm()
    errors: list[str] = []
    created = 0

    numeric_fields = [
        "filament_price_per_kg",
        "filament_weight_g",
        "print_time_hours",
        "labour_time_minutes",
        "margin_percentage",
    ]

    def process_payload(raw_payload, line_no):
        nonlocal created
        cleaned = {
            "piece_name": (raw_payload.get("piece_name") or "").strip(),
        }
        piece_name = cleaned["piece_name"]
        if piece_name:
            exists_qs = PrintJob.objects.filter(user=request.user, name__iexact=piece_name)
            if exists_qs.exists():
                raise ValueError("Ja existe uma pe\u00e7a com este nome.")
        for key in numeric_fields:
            try:
                cleaned[key] = parse_decimal(raw_payload.get(key))
            except ValueError as exc:
                raise ValueError(f"{key}: {exc}") from exc

        if cleaned["margin_percentage"] >= Decimal("100"):
            raise ValueError("Margem deve ser inferior a 100%.")

        result = calculate_print_job(cleaned)

        PrintJob.objects.create(
            user=request.user,
            name=cleaned["piece_name"],
            filament_price_per_kg=cleaned["filament_price_per_kg"],
            filament_weight_g=cleaned["filament_weight_g"],
            print_time_hours=cleaned["print_time_hours"],
            labour_time_minutes=cleaned["labour_time_minutes"],
            margin_percentage=cleaned["margin_percentage"],
            cost_filament=result["cost_filament"],
            cost_energy=result["cost_energy"],
            cost_labour=result["cost_labour"],
            cost_machine=result["cost_machine"],
            cost_total=result["cost_total"],
            price_final=result["price_final"],
            consumption_kwh=result["consumption_kwh"],
        )
        created += 1

    if request.method == "POST":
        form = PieceImportForm(request.POST, request.FILES)
        if form.is_valid():
            uploaded = form.cleaned_data["file"]
            ext = Path(uploaded.name or "").suffix.lower()

            if ext != ".xlsx":
                errors.append("Formato nAo suportado. Utilize um ficheiro Excel (.xlsx).")
            else:
                try:
                    from openpyxl import load_workbook  # type: ignore
                except ImportError:
                    errors.append(
                        "Suporte a Excel indisponAvel (biblioteca openpyxl nAo instalada)."
                    )
                else:
                    try:
                        uploaded.seek(0)
                        workbook = load_workbook(uploaded, data_only=True)
                        sheet = workbook.active
                    except Exception as exc:  # noqa: BLE001
                        errors.append(f"Erro ao ler Excel: {exc}")
                    else:
                        rows = list(sheet.iter_rows(values_only=True))
                        if not rows:
                            errors.append("O ficheiro Excel estA vazio.")
                        else:
                            headers = [
                                (str(cell).strip() if cell is not None else "")
                                for cell in rows[0]
                            ]
                            missing = [
                                col for col in IMPORT_COLUMNS if col not in headers
                            ]
                            if missing:
                                errors.append(
                                    "Colunas em falta: " + ", ".join(missing)
                                )
                            else:
                                index_map = {name: headers.index(name) for name in IMPORT_COLUMNS}
                                for row_number, row in enumerate(rows[1:], start=2):
                                    payload = {}
                                    for key, idx in index_map.items():
                                        payload[key] = row[idx] if idx < len(row) else None
                                    try:
                                        process_payload(payload, row_number)
                                    except Exception as exc:  # noqa: BLE001
                                        errors.append(f"Linha {row_number}: {exc}")

            if created:
                messages.success(
                    request,
                    f"Importadas {created} peAa(s) para o seu utilizador.",
                )
                if not errors:
                    return redirect("pieces_list")

    return render(
        request,
        "core/piece_import.html",
        {
            "form": form,
            "errors": errors,
            "expected_columns": IMPORT_COLUMNS,
        },
    )









